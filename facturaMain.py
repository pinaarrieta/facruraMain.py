#****************************************  Extractor XML   *******************************************************
#EricK Genaro Piña Arrieta Barbosa
#Proyecto: Extractor de datos XML en Phyton
#Fecha: 31/10/2025
#Descripción: Este proyecto tiene como objetivo extraer y procesar datos de un XML CFDI 4 utilizando Python.
#*****************************************************************************************************************

#1. Importar bibliotecas necesarias ******************************************************************************
#*****************************************************************************************************************

import pandas as pd
import xml.etree.ElementTree as ET #Para leer los elementos de un XML
from openpyxl import load_workbook #Para escribir archivos excel xlsx

#2.Definir la ruta de los archivo XML almacenados enla PC del usuario ********************************************
#*****************************************************************************************************************

def definir_ruta_archivos(r):
    print('\n')
    print('Probablente tus archivos se encuentran en la carpeta de Descargas')
    print('Ejemplo de ruta en Windows: C:/Users/TuUsuario/Downloads/CFDI00000000.xml')
    print('Ejemplo de ruta en Linux Ubuntu: /home/TuUsuario/Descargas/CFDI00000000.xml')
    print('No olvides anotar el .xml de tu archivo')
    print('\n')
    ruta = input('Ingrese la ruta donde se encuentran los archivos XML:')
    print('\n')
    print('Verica si la ruta y el nombre del archivo son correctos:')
    input('Presiona ENTER para continuar ...')
    print('\n') 
    return ruta

#3.Definir procedimiento para parsear un CFDI 4.00, extraer los datos relevantes y mostrarlos en pantalla*********
#*****************************************************************************************************************

def parseXML(p):
    # Ruta del archivo XML del CFDI, asegurarse de antes llamar a definir_ruta_archivos() desde el main
    xml_file = definir_ruta_archivos(0)

    # Definir el espacio de nombres (namespace), obligatorio en el CFDI 4.0
    namespaces = {'cfdi': 'http://www.sat.gob.mx/cfd/4'}
        #Para consulta de todos los nodos, atributos y datos del CFDI 4.0, 
        # ver http://www.sat.gob.mx/sitio_internet/cfd/4/cfdv40.xsd
        # ver también http://omawww.sat.gob.mx/tramitesyservicios/Paginas/anexo_20.htm
        # Solo exploramos el nodo raiz y los nodos emisor, receptor e impuestos del 1er nivel del árbol

    # Cargar y parsear (leer y desmenuzar) el archivo XML, identifica etiquetas, atributos y valores
    tree = ET.parse(xml_file)
    root = tree.getroot() # Nodo raíz del XML

    #Extraer datos generales del CFDI, que estan en el nodo raiz del XML
    serie = root.attrib.get('Serie') #Las etiquetas son fijas para todos los xml v4
    folio = root.attrib.get('Folio')
    fecha = root.attrib.get('Fecha')
    subtotal = root.attrib.get('SubTotal')
    descuento = root.attrib.get('Descuento')
    total = root.attrib.get('Total')

    # Buscar el nodo <cfdi:Emisor> para luego acceder a sus atributos y valores
    emisor = root.find('cfdi:Emisor', namespaces)

    # Buscar el nodo <cfdi:Receptor>
    receptor = root.find('cfdi:Receptor', namespaces)

    #Buscar el nodo <cfdi:Impuestos>
    impuestos = root.find('cfdi:Impuestos', namespaces)

    #Extraer e imprimir los atributos y valores del nodo raiz, emisor, receptor e impuestos 
    #un datos del nodo Impuestos y otro del nodo Complemento
    print('Estos son los datos de tu CFDI  y XML, revisalos:\n')
    print("Serie:", serie)
    print("Folio:", folio)
    print("Fecha:", fecha)
    print("SubTotal:", subtotal)
    print("Descuento:", descuento)#Opcional, puede no existir, en ese caso devuelve None

    if impuestos is not None: #este dato es del nodo Impuestos, 
            #is not None es necesario ya que puede ser que no tenga impuestos o esten exentos, 
            #en cuyo caso devolverá None
        print("Total_Impuestos_trasladados:", impuestos.attrib.get('TotalImpuestosTrasladados'))
        totalImpuestosTrasladados = impuestos.attrib.get('TotalImpuestosTrasladados') #Variable para el return 
            #Este impuesto es el IVA, pero también incluye el IEPS si es que aplica
    print("Total:", total)

    #Extraer e imprimir los atributos y valores del nodo Emisor
    if emisor is not None:
        print("RFC_Emisor:", emisor.attrib.get('Rfc'))
        print("Nombre_Emisor:", emisor.attrib.get('Nombre'))
        RFC_Emisor = emisor.attrib.get('Rfc')  #Variable para el return
        Nombre_Emisor = emisor.attrib.get('Nombre')  #Variable para el return
    #Extraer e imprimir los atributos y valores del nodo Receptor
    if receptor is not None:
        print("RFC_Receptor:", receptor.attrib.get('Rfc'))
        print("Nombre_Receptor:", receptor.attrib.get('Nombre'))
        RFC_Receptor = receptor.attrib.get('Rfc')  #Variable para el return
        Nombre_Receptor = receptor.attrib.get('Nombre')  #Variable para el return

    #Elaborar rutinas para datos de moneda que deberan pasarse por float(), 
    #pero como son opcionales, podrán ser None (string), si es un string debe convertirse a 0.0
    #los datos opcionales son descuento y totalImpuestosTrasladados
    if descuento is None:
        descuento = 0.0 #Así el usuario puede usar su excel para sumatorias
    if totalImpuestosTrasladados is None:
        totalImpuestosTrasladados = 0.0
        
    miFila = [serie, folio, fecha, float(subtotal), float(descuento), float(totalImpuestosTrasladados), 
                  float(total), RFC_Emisor, Nombre_Emisor, RFC_Receptor, Nombre_Receptor]
        
    return(miFila)  #Regresa la lista con los datos extraidos del XML

#4. Crear el archivo en excel Vacio con encabezados
#*****************************************************************************************************************

def crearExcelconEncabezados(nombreXLS):
    #Definir la lista de encabezados
    encabezados = {'Serie': [],
                   'Folio': [],
                   'Fecha' : [],
                   'SubTotal': [],
                   'Descuento' : [],
                   'Total_Impuestos_trasladados': [],
                   'Total': [],
                   'RFC_Emisor': [], 
                   'Nombre_Emisor': [],
                   'RFC_Receptor': [],
                   'Nombre_Receptor': [],}

    #Crear un dataframe vacio con los encabezados
    df = pd.DataFrame(columns=encabezados)

    #Vista previas del dataframe
    print('\n')
    print('Vista previa del dataframe: ')
    print(df)

    #Definir la ruta donde se guardará el archivo excel
    rutaXLS = input('\nDefine solo la ruta donde se guardará el archivo: \n')
    

    #Crear el archivo excel vacio con encabezados
    df.to_excel(rutaXLS + nombreXLS, index=False)

    return rutaXLS + nombreXLS

#5. Convertir una lista (registro) en un diccionario con los encabezados de excel antes definidos
#*****************************************************************************************************************

def convertir_lista_a_diccionario(registro):
    #Esta lista es el return de parseXML(0), y hay que convertirla en un diccionarios con los encabezados de excel
    #Solo hay que pasarle el parametro 'registro' que es return de parseXML(0)

    listaParaDiccionario = registro

    #A cada dato le asignamos una variable
    se, fo, fe, su, de, tit, to, re, ne, rr, nr = listaParaDiccionario

    #Crear el diccionario

    diccionario ={'Serie' : se,
                'Folio' : fo,
                'Fecha'	: fe,
                'SubTotal' : su,	
                'Descuento' : de,
                'Total_Impuestos_trasladados' : tit,
                'Total' : to,
                'RFC_Emisor' : re,
                'Nombre_Emisor' : ne,
                'RFC_Receptor' : rr,
                'Nombre_Receptor' : nr,
                }
    return(diccionario)

#6. Agregar la los datos del CFDI en el archivo de excel
#*****************************************************************************************************************

def insertarFilaCFDI(ruta_archivo):
    
    #Definir el diccionario para el dataframe
    nueva_fila = elDiccionarioListo

    # Convertimos el diccionario anterior en un DataFrame
    df_nueva = pd.DataFrame([nueva_fila])

    # Tomamos nuestro archivo en Excel que ya tiene sus encabezados en la fila 1
    libro = load_workbook(ruta_archivo)
    nombre_hoja = libro.sheetnames[0] #Aqui le indicamos que tome la primera hoja del excel

    # Leemos el dataframe
    hoja_existente = pd.read_excel(ruta_archivo, sheet_name = nombre_hoja)

    # Escribimos el la siguiente fila
    with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                                                        #El modo a evita que borremos lo que ya esta en las celdas
        startrow = len(hoja_existente) + 1  # Siguiente fila libre
        df_nueva.to_excel(writer, sheet_name=nombre_hoja, index=False, header=False, startrow=startrow)

    print('\n')
    print("Fila agregada exitosamente .")
    print('\n')
    
#inicialización del programa**************************************************************************************
if __name__ == "__main__":
    #1. Proemio de la aplicación**********************************************************************************
    print('\n')
    print('**********************************************************************************')
    print('**************************Extractor de datos XML CFDI 4.0*************************')
    print('**********************************************************************************')
    print('\n')

    #2. Llamar al procedimiento de creación de un archivo en excel************************************************
    print('A continuación deberas indicar  por separado un nombre y una ruta para tu arcvhivo excel:')
    print('Ejemplo de ruta en Windows: C:/Users/TuUsuario/Downloads/')  
    print('Ejemplo de ruta en Linux Ubuntu: /home/TuUsuario/Descargas/\n')
    print('Ejemplo de nombre del Excel: miArchivo  (sin extensión)')
    nombreXLS = input('Define solo el nombre de tu archivo Excel, sin extensión: \n')
    nombreXLS = nombreXLS + '.xlsx' #Este será el parámetros de la funcion siguiente

    pathXLS = crearExcelconEncabezados(nombreXLS)
    print('\n')
    print('El archivo fue creado correctamente en: \n', pathXLS, '\n',)
    input('Ahora pasaremos los datos de XML a tu archivo en excel, presiona Enter para continuar')
    rutaArchivo = pathXLS

    #3. Iniciamos un bucle a fin de cargar CFDI por CFDI, es necesario que el usuario corroboré los datos
    print('\n')
    print('Ubica bien la ruta de la carpeta(s) y nombre(s) de tur archivos XML')
    
    si_no = input('¿Estás listo para cargar tus CFDI-XML?, s/n: ')

    while si_no == 's' or si_no == 'S':
        #3.1 Lama al procedimiento para leer XML
        registro = (parseXML(0)) 
        print('\n')
        print('ATENCIÓN: Verifica si los datos del CFDI son correctos') 
        input("Presiona ENTER para continuar ...") 
        print('\n')
        print('Los datos que se agregarán en la fila de excel son: \n', registro)
        print('\n')

        #3.2 Llamar al procedimiento que convierte la lista (registro) en un diccionario con los encabezados de excel
        print('\n')
        print(registro)
        elDiccionarioListo = convertir_lista_a_diccionario(registro)
        print('\n')
        print('El diccionario fue creado satisfactoriamente:\n')
        print('Esta es su estructura con sus encabezados: \n', elDiccionarioListo )
        input('\nPresiona ENTER para continuar')

        #3.3  Llamar al procedimiento que insertará el registro en el archivo de excel
        insertarFilaCFDI(pathXLS)

        #3.4 Para salir o seguir en el bucle
        si_no = input('¿Deseas agregar más CFDI?: s/n ')

    #Pie de página de la aplicación y fin del programa
    print('\n')
    print('*************************************************************************************************************')
    print('**************************************Fin del programa*******************************************************')
    print('*************************************************************************************************************')
    print('Gracias por usar la aplicación, cualquier comentario a pinaarrieta@yahoo.com.mx\n')
    input('Presiona ENTER para finalizar....')
